import openpyxl
from openpyxl import Workbook
import os

folder_path = input('Введите полный путь к папке с файлами .xlsx - ')  # указываем путь к папке с файлами .xlsx
hat = str(input('Введите шапку - '))
save_path = str(input('Введите путь для сохранения - '))
result_filename = str(input("Введите название файла (БЕЗ РАСШИРЕНИЯ), под которым он будет сохранен - "))
files = os.listdir(folder_path)  # получаем список файлов в папке
result = openpyxl.Workbook()  # создаем новый файл
result.create_sheet("Результат", 0)
res_shit = result.active
cnt = 0

for file in files:

    if file.endswith(".xlsx"):  # проверяем, является ли файл файлом формата .xlsx

        workbook = openpyxl.load_workbook(folder_path + '\\' + file)  # открываем активный файл
        first_row = [cell.value for cell in workbook.active[1]]
        if hat in first_row:
            cell_value = workbook['Лист1']['D2'].value  # получаем значение ячейки D2
            range_of_rows = workbook['Лист1'][9:workbook['Лист1'].max_row]  # получаем диапазон строк, начиная со строчки 9, в которых есть значения
            res_shit.append([cell_value])  # записываем значение ячейки D2

            # записываем строки от 9 до первой пустой строки
            for row in range_of_rows:
                if all([cell.value is None for cell in row]):
                    break
                res_shit.append([cell.value for cell in row])

        else:
            print("Шапка не была найдена в первой строчке")
    else:
        print(f"Файл {file} не соответствует формату .xlsx")

# сохраняем новый файл в папку save_path
result.save(save_path + "\\" + result_filename + '.xlsx')
print("Готово!")